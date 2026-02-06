import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import warnings
warnings.filterwarnings('ignore')

# Configuration de la page
st.set_page_config(
    page_title="Générateur de Tableaux Scolaires",
    page_icon=":bar_chart:",
    layout="wide"
)

# Style CSS personnalisé
st.markdown("""
<style>
    .main-header {
        color: #FFFFFF;
        text-align: center;
        padding: 2rem;
        background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%);
        border-radius: 10px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .stButton>button {
        background: linear-gradient(135deg, #3B82F6 0%, #1E3A8A 100%);
        color: white;
        border: none;
        padding: 0.75rem 2rem;
        border-radius: 25px;
        font-weight: bold;
        transition: all 0.3s ease;
        width: 100%;
        font-size: 1.1rem;
    }
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 5px 15px rgba(59, 130, 246, 0.4);
    }
    .upload-section {
        background: #f0f9ff;
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        border: 2px dashed #3B82F6;
    }
    .success-message {
        background: #dcfce7;
        color: #166534;
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #10b981;
        margin: 1rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .info-box {
        background: #e0f2fe;
        color: #075985;
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #0ea5e9;
        margin: 1rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .warning-box {
        background: #fef3c7;
        color: #92400e;
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #f59e0b;
        margin: 1rem 0;
    }
    .stat-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        text-align: center;
        border-top: 4px solid #3B82F6;
    }
    .graph-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        margin-bottom: 1.5rem;
    }
</style>
""", unsafe_allow_html=True)

# En-tête de l'application
st.markdown("""
<div class="main-header">
    <h1 style="margin: 0; font-size: 2.5rem;">📊 Générateur de Tableaux et Statistiques Scolaires</h1>
    <p style="margin-top: 1rem; font-size: 1.2rem; opacity: 0.9;">Importez votre base de données Excel et générez automatiquement les tableaux par année + statistiques</p>
</div>
""", unsafe_allow_html=True)

def apply_excel_formatting(writer, sheet_name, df):
    """Applique le formatage Excel pour une feuille donnée"""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Définir les styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Formater l'en-tête
    for col_num, value in enumerate(df.columns.values, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border
    
    # Formater les données
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            # Aligner à gauche pour les colonnes de texte, à centre pour les nombres
            if col in [1, 2, 3]:  # École, Moughataa, Commune
                cell.alignment = align_left
            else:
                cell.alignment = align_center
            cell.border = border
    
    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 30,  # École
        'B': 15,  # Moughataa
        'C': 20,  # Commune
        'D': 12,  # Nbre d'élèves
        'E': 15,  # Nbre d'enseignants
        'F': 12,  # Nbre de DP
        'G': 20,  # Ratio moyen élèves/DP
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width

def create_annual_tables(df):
    """Crée les tableaux pour chaque année à partir du format spécifique"""
    
    # Vérifier la structure du fichier
    expected_columns = 21  # 3 colonnes de base + 6*3 colonnes par année
    if len(df.columns) != expected_columns:
        st.error(f"❌ Structure de fichier incorrecte. Attendu: {expected_columns} colonnes, obtenu: {len(df.columns)}")
        return None
    
    # Initialiser le writer Excel
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    # Pour chaque année de 1 à 6
    for year in range(1, 7):
        year_str = f"Année {year}"
        sheet_name = f"Année_{year}"
        
        # Calculer les indices de colonnes pour cette année
        start_idx = 3 + (year - 1) * 3  # 3 colonnes de base
        
        # Vérifier si les indices sont valides
        if start_idx + 2 >= len(df.columns):
            st.error(f"❌ Indice de colonne invalide pour l'année {year}")
            continue
        
        # Sélectionner les colonnes pour cette année
        base_cols = ['Region', 'Moughataa', "Nom de l'ecole"]
        
        # Colonnes pour l'année spécifique
        eleves_col = df.columns[start_idx + 2]  # Nbre Eleves
        enseign_col = df.columns[start_idx + 1]  # Nbre enseign
        dp_col = df.columns[start_idx]  # Nbre DP
        
        year_cols = base_cols + [eleves_col, enseign_col, dp_col]
        
        # Créer le DataFrame pour cette année
        year_df = df[year_cols].copy()
        
        # Renommer les colonnes selon l'image
        rename_dict = {
            'Region': 'Commune',
            'Moughataa': 'Moughataa',
            "Nom de l'ecole": 'École',
            eleves_col: 'Nbre d\'élèves',
            enseign_col: 'Nbre d\'enseignants',
            dp_col: 'Nbre de DP'
        }
        
        year_df = year_df.rename(columns=rename_dict)
        
        # Réorganiser les colonnes
        year_df = year_df[['École', 'Moughataa', 'Commune', 
                          'Nbre d\'élèves', 'Nbre d\'enseignants', 'Nbre de DP']]
        
        # CORRECTION: Calculer le ratio moyen élèves/DP (élèves ÷ DP)
        year_df['Ratio moyen élèves/DP'] = year_df.apply(
            lambda row: round(row['Nbre d\'élèves'] / row['Nbre de DP'], 1) 
            if pd.notnull(row['Nbre de DP']) and row['Nbre de DP'] > 0 else 0,
            axis=1
        )
        
        # Trier par Moughataa puis par nom d'école
        year_df = year_df.sort_values(['Moughataa', 'École'])
        
        # Ajouter des totaux en bas
        total_eleves = year_df['Nbre d\'élèves'].sum()
        total_enseignants = year_df['Nbre d\'enseignants'].sum()
        total_dp = year_df['Nbre de DP'].sum()
        total_ratio = round(total_eleves / total_dp, 1) if total_dp > 0 else 0
        
        total_row = pd.DataFrame({
            'École': ['TOTAL'],
            'Moughataa': [''],
            'Commune': [''],
            'Nbre d\'élèves': [total_eleves],
            'Nbre d\'enseignants': [total_enseignants],
            'Nbre de DP': [total_dp],
            'Ratio moyen élèves/DP': [total_ratio]
        })
        
        year_df_with_totals = pd.concat([year_df, total_row], ignore_index=True)
        
        # Ajouter à l'Excel
        year_df_with_totals.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Appliquer le formatage
        apply_excel_formatting(writer, sheet_name, year_df_with_totals)
        
        # Formater la ligne de totaux
        worksheet = writer.sheets[sheet_name]
        last_row = len(year_df_with_totals) + 1
        
        # Mettre en gras la ligne de totaux
        for col in range(1, 8):  # 7 colonnes
            cell = worksheet.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            if col >= 4:  # Les colonnes numériques (4-7)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Sauvegarder
    writer.close()
    output.seek(0)
    
    return output

def create_statistical_graphs(df, selected_year, x_variable, y_variable, graph_type):
    """Crée des graphiques statistiques binaires"""
    
    # Préparer les données pour l'année sélectionnée
    start_idx = 3 + (selected_year - 1) * 3
    
    if start_idx + 2 >= len(df.columns):
        st.error("❌ Données non disponibles pour l'année sélectionnée")
        return None, None
    
    # Créer un DataFrame pour l'analyse
    analysis_df = df[['Region', 'Moughataa', "Nom de l'ecole"]].copy()
    
    # Ajouter les données pour l'année sélectionnée
    analysis_df['Nbre DP'] = df.iloc[:, start_idx]
    analysis_df['Nbre enseign'] = df.iloc[:, start_idx + 1]
    analysis_df['Nbre Eleves'] = df.iloc[:, start_idx + 2]
    
    # Calculer le ratio élèves/DP
    analysis_df['Ratio élèves/DP'] = analysis_df.apply(
        lambda row: round(row['Nbre Eleves'] / row['Nbre DP'], 1) 
        if pd.notnull(row['Nbre DP']) and row['Nbre DP'] > 0 else 0,
        axis=1
    )
    
    # Renommer les colonnes pour l'affichage
    analysis_df = analysis_df.rename(columns={
        'Region': 'Commune',
        "Nom de l'ecole": 'École'
    })
    
    # Créer le graphique en fonction du type sélectionné
    fig = None
    
    if graph_type == "Nuage de points":
        fig = px.scatter(
            analysis_df, 
            x=x_variable, 
            y=y_variable,
            hover_data=['École', 'Moughataa', 'Commune'],
            title=f"Nuage de points: {x_variable} vs {y_variable} (Année {selected_year})",
            labels={x_variable: x_variable, y_variable: y_variable},
            color='Moughataa'
        )
        
    elif graph_type == "Histogramme":
        fig = px.histogram(
            analysis_df, 
            x=x_variable,
            title=f"Distribution de {x_variable} (Année {selected_year})",
            labels={x_variable: x_variable, 'count': 'Nombre d\'écoles'},
            color='Moughataa',
            nbins=20
        )
        
    elif graph_type == "Diagramme en barres":
        # Regrouper par Moughataa
        grouped_data = analysis_df.groupby('Moughataa').agg({
            x_variable: 'mean',
            y_variable: 'mean'
        }).reset_index()
        
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        
        fig.add_trace(
            go.Bar(
                x=grouped_data['Moughataa'],
                y=grouped_data[x_variable],
                name=x_variable,
                marker_color='#3B82F6'
            ),
            secondary_y=False
        )
        
        fig.add_trace(
            go.Scatter(
                x=grouped_data['Moughataa'],
                y=grouped_data[y_variable],
                name=y_variable,
                mode='lines+markers',
                marker_color='#EF4444',
                line=dict(width=3)
            ),
            secondary_y=True
        )
        
        fig.update_layout(
            title=f"Moyenne de {x_variable} et {y_variable} par Moughataa (Année {selected_year})",
            xaxis_title="Moughataa",
            showlegend=True
        )
        
        fig.update_yaxes(title_text=x_variable, secondary_y=False)
        fig.update_yaxes(title_text=y_variable, secondary_y=True)
        
    elif graph_type == "Box plot":
        fig = px.box(
            analysis_df, 
            x='Moughataa', 
            y=x_variable,
            title=f"Distribution de {x_variable} par Moughataa (Année {selected_year})",
            points="all",
            hover_data=['École', 'Commune']
        )
    
    elif graph_type == "Carte thermique (heatmap)":
        # Créer une matrice de corrélation
        numeric_cols = ['Nbre DP', 'Nbre enseign', 'Nbre Eleves', 'Ratio élèves/DP']
        corr_matrix = analysis_df[numeric_cols].corr()
        
        fig = px.imshow(
            corr_matrix,
            text_auto=True,
            title=f"Matrice de corrélation (Année {selected_year})",
            color_continuous_scale='RdBu',
            aspect="auto"
        )
    
    if fig:
        # Personnaliser le graphique
        fig.update_layout(
            template="plotly_white",
            hovermode="closest",
            height=500,
            font=dict(size=12)
        )
    
    return fig, analysis_df

def main():
    # Section de téléversement
    st.markdown('<div class="upload-section">', unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "Téléversez votre fichier Excel (BD.xlsx)",
        type=['xlsx', 'xls'],
        help="Le fichier doit avoir exactement 21 colonnes"
    )
    st.markdown('</div>', unsafe_allow_html=True)
    
    if uploaded_file is not None:
        try:
            # Lire le fichier Excel
            df = pd.read_excel(uploaded_file, engine='openpyxl')
            
            # Afficher les informations sur la structure
            st.markdown('<div class="info-box">', unsafe_allow_html=True)
            st.success("✅ Fichier téléversé avec succès !")
            
            # Afficher la structure détectée
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown("""
                <div class="stat-card">
                    <h3>🏫</h3>
                    <h2>{}</h2>
                    <p>Écoles</p>
                </div>
                """.format(len(df)), unsafe_allow_html=True)
            
            with col2:
                st.markdown("""
                <div class="stat-card">
                    <h3>📊</h3>
                    <h2>{}</h2>
                    <p>Colonnes</p>
                </div>
                """.format(len(df.columns)), unsafe_allow_html=True)
            
            with col3:
                st.markdown("""
                <div class="stat-card">
                    <h3>📅</h3>
                    <h2>6</h2>
                    <p>Années</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                st.markdown("""
                <div class="stat-card">
                    <h3>🔢</h3>
                    <p>3 colonnes × 6 ans</p>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
            
            # Créer des onglets pour différentes fonctionnalités
            tab1, tab2 = st.tabs(["📋 Génération des Tableaux", "📈 Analyse Statistique"])
            
            with tab1:
                # Bouton pour générer les tableaux
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    generate_button = st.button(
                        "🚀 Générer les 6 Tableaux par Année",
                        type="primary",
                        use_container_width=True
                    )
                
                if generate_button:
                    with st.spinner("🔄 Génération des tableaux en cours..."):
                        try:
                            # Vérifier la structure
                            if len(df.columns) != 21:
                                st.error(f"❌ Le fichier doit avoir exactement 21 colonnes. Votre fichier en a {len(df.columns)}.")
                            else:
                                # Créer les tableaux
                                excel_output = create_annual_tables(df)
                                
                                if excel_output:
                                    # Afficher le message de succès
                                    st.markdown('<div class="success-message">', unsafe_allow_html=True)
                                    st.success("🎉 Tableaux générés avec succès !")
                                    
                                    st.markdown("""
                                    ### ✅ Génération terminée
                                    
                                    **Structure des tableaux générés :**
                                    1. **École** (Nom de l'école)
                                    2. **Moughataa**
                                    3. **Commune** (Region)
                                    4. **Nbre d'élèves**
                                    5. **Nbre d'enseignants**
                                    6. **Nbre de DP**
                                    7. **Ratio moyen élèves/DP** (calculé: élèves ÷ DP)
                                    
                                    **Caractéristiques :**
                                    - 📑 6 feuilles Excel (Année_1 à Année_6)
                                    - 🧮 Ratio calculé automatiquement
                                    - 📊 Ligne de totaux avec mise en forme
                                    - 🎨 Formatage professionnel
                                    - 🔍 Tri par Moughataa et École
                                    """)
                                    st.markdown('</div>', unsafe_allow_html=True)
                                    
                                    # Bouton de téléchargement
                                    st.download_button(
                                        label="📥 Télécharger le fichier Excel complet",
                                        data=excel_output,
                                        file_name="Tableaux_Scolaires_Par_Annee.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        type="primary",
                                        use_container_width=True
                                    )
                                    
                                else:
                                    st.markdown('<div class="warning-box">', unsafe_allow_html=True)
                                    st.warning("⚠️ Impossible de générer les tableaux.")
                                    st.markdown('</div>', unsafe_allow_html=True)
                            
                        except Exception as e:
                            st.error(f"❌ Erreur lors de la génération : {str(e)}")
            
            with tab2:
                st.markdown("## 📈 Analyse Statistique Binaire")
                st.markdown("Générez des graphiques pour analyser les relations entre différentes variables.")
                
                # Section de sélection des paramètres
                with st.container():
                    st.markdown("### ⚙️ Paramètres de l'analyse")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        # Sélection de l'année
                        selected_year = st.selectbox(
                            "Sélectionnez l'année:",
                            options=[1, 2, 3, 4, 5, 6],
                            format_func=lambda x: f"Année {x}"
                        )
                    
                    with col2:
                        # Sélection du type de graphique
                        graph_type = st.selectbox(
                            "Type de graphique:",
                            options=["Nuage de points", "Histogramme", "Diagramme en barres", 
                                    "Box plot", "Carte thermique (heatmap)"]
                        )
                    
                    with col3:
                        # Variables disponibles pour l'analyse
                        available_variables = [
                            "Nbre DP", "Nbre enseign", "Nbre Eleves", "Ratio élèves/DP"
                        ]
                        
                        if graph_type in ["Nuage de points", "Diagramme en barres"]:
                            # Pour ces graphiques, besoin de deux variables
                            x_variable = st.selectbox(
                                "Variable X:",
                                options=available_variables,
                                index=0
                            )
                            y_variable = st.selectbox(
                                "Variable Y:",
                                options=available_variables,
                                index=1
                            )
                        else:
                            # Pour les autres graphiques, une seule variable suffit
                            x_variable = st.selectbox(
                                "Variable à analyser:",
                                options=available_variables,
                                index=0
                            )
                            y_variable = x_variable  # Même variable pour Y
                    
                    # Bouton pour générer le graphique
                    generate_graph_button = st.button(
                        "📊 Générer le graphique",
                        type="primary",
                        use_container_width=True
                    )
                
                # Section d'affichage des résultats
                if generate_graph_button:
                    with st.spinner("🔄 Génération du graphique en cours..."):
                        try:
                            # Créer le graphique
                            fig, analysis_df = create_statistical_graphs(
                                df, selected_year, x_variable, y_variable, graph_type
                            )
                            
                            if fig:
                                # Afficher le graphique
                                st.markdown('<div class="graph-card">', unsafe_allow_html=True)
                                st.plotly_chart(fig, use_container_width=True)
                                st.markdown('</div>', unsafe_allow_html=True)
                                
                                # Afficher les statistiques résumées
                                st.markdown("### 📊 Statistiques descriptives")
                                
                                col1, col2, col3, col4 = st.columns(4)
                                
                                with col1:
                                    st.metric(
                                        "Moyenne élèves/DP",
                                        f"{analysis_df['Ratio élèves/DP'].mean():.1f}"
                                    )
                                
                                with col2:
                                    st.metric(
                                        "Total élèves",
                                        f"{analysis_df['Nbre Eleves'].sum():,}"
                                    )
                                
                                with col3:
                                    st.metric(
                                        "Total enseignants",
                                        f"{analysis_df['Nbre enseign'].sum():,}"
                                    )
                                
                                with col4:
                                    st.metric(
                                        "Total DP",
                                        f"{analysis_df['Nbre DP'].sum():,}"
                                    )
                                
                                # Afficher un aperçu des données d'analyse
                                with st.expander("🔍 Voir les données d'analyse", expanded=False):
                                    st.dataframe(
                                        analysis_df.sort_values('Ratio élèves/DP', ascending=False),
                                        use_container_width=True
                                    )
                                    
                                    # Option pour télécharger les données d'analyse
                                    csv_data = analysis_df.to_csv(index=False).encode('utf-8')
                                    st.download_button(
                                        label="📥 Télécharger les données d'analyse (CSV)",
                                        data=csv_data,
                                        file_name=f"donnees_analyse_annee_{selected_year}.csv",
                                        mime="text/csv"
                                    )
                            
                        except Exception as e:
                            st.error(f"❌ Erreur lors de la génération du graphique : {str(e)}")
                
                # Section d'exemples de graphiques
                with st.expander("💡 Exemples d'analyses possibles", expanded=False):
                    st.markdown("""
                    **Exemples d'analyses binaires intéressantes :**
                    
                    1. **Élèves vs Enseignants** : Relation entre le nombre d'élèves et le nombre d'enseignants
                    2. **Ratio vs DP** : Comment le ratio élèves/DP varie avec le nombre de DP
                    3. **Distribution par Moughataa** : Comparer les moyennes entre différentes régions
                    4. **Corrélations** : Identifier les relations entre différentes variables
                    
                    **Conseils :**
                    - Utilisez le nuage de points pour identifier des tendances
                    - L'histogramme montre la distribution d'une variable
                    - Le box plot permet de comparer les distributions entre régions
                    - La carte thermique révèle les corrélations entre toutes les variables
                    """)
        
        except Exception as e:
            st.error(f"❌ Erreur lors de la lecture du fichier : {str(e)}")
            st.markdown('<div class="warning-box">', unsafe_allow_html=True)
            st.info("""
            **Conseils de dépannage :**
            1. Vérifiez que le fichier est bien un Excel (.xlsx ou .xls)
            2. Ouvrez le fichier dans Excel pour vérifier sa structure
            3. Assurez-vous que les colonnes sont dans le bon ordre
            """)
            st.markdown('</div>', unsafe_allow_html=True)
    
    else:
        # Afficher un exemple de structure attendue
        st.info("👆 Veuillez téléverser un fichier Excel pour commencer")
        
        # Exemple de structure
        with st.expander("🧾 Structure exacte attendue", expanded=True):
            st.markdown("""
            **Votre fichier doit avoir exactement 21 colonnes dans cet ordre :**
            
            1. **Region**
            2. **Moughataa**
            3. **Nom de l'ecole**
            4. Nbre DP (Année 1)
            5. Nbre enseign (Année 1)
            6. Nbre Eleves (Année 1)
            7. Nbre DP (Année 2)
            8. Nbre enseign (Année 2)
            9. Nbre Eleves (Année 2)
            10. Nbre DP (Année 3)
            11. Nbre enseign (Année 3)
            12. Nbre Eleves (Année 3)
            13. Nbre DP (Année 4)
            14. Nbre enseign (Année 4)
            15. Nbre Eleves (Année 4)
            16. Nbre DP (Année 5)
            17. Nbre enseign (Année 5)
            18. Nbre Eleves (Année 5)
            19. Nbre DP (Année 6)
            20. Nbre enseign (Année 6)
            21. Nbre Eleves (Année 6)
            
            **Note :** Le ratio est calculé comme : **Nbre d'élèves ÷ Nbre de DP**
            """)

if __name__ == "__main__":
    main()